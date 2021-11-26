# coding=utf-8
#excel表格操作
import openpyxl
import datetime
import os
from openpyxl import worksheet
from openpyxl import load_workbook

#excel读取类
class excelProcess:
    sheet:worksheet = None   #当前读取的表格
    sheet_name:str = ""  #工作表名
    max_row:int = 0   #最大行数
    max_col:int = 0    #最大列数
    curr_row:int = 0    #当前使用的行数
    curr_col:int  = 0   #当前使用的列数
    __excel_op:openpyxl.Workbook = None   #读取的表格文件实体
    __excel_src:str = ""  #表格路径

    #读取表格
    #file_src ->文件路径  sheet_name -> 表格名
    def readExcel(self, file_src:str, sheet_name:str):
        self.__excel_op = openpyxl.load_workbook(file_src)  #打开excel
        if self.__excel_op != None:
            self.__excel_src = file_src
            self.sheet = self.__excel_op[sheet_name]
            if self.sheet != None:
                #更新最大行与列
                self.max_row = self.sheet.max_row
                self.max_col = self.sheet.max_column
                self.sheet_name = sheet_name
        self.__excel_op.close()

    #获取某一行数据,并以列表方式返回
    def getTableRowValue(self, row:int)->list:
        data_tu = []
        if self.sheet == None:
            print("无表格数据！")
        else:
            #读取列
            for inx in range(1, self.max_col):
                line = self.sheet.cell(row, inx)
                if line != "" or line != None:
                    data_tu.append(line.value)
        return data_tu

    #获取某一列数据,并以列表方式返回
    def getTableColumValue(self,col:int)->list:
        data_tu = []
        if self.sheet == None:
            print("无表格数据！")
        else:
            #读取行
            for inx in range(1, self.max_col):
                colum = self.sheet.cell(inx, col)
                if colum != "" or colum != None:
                    data_tu.append(colum.value)
        return data_tu

    #写入数据到表格
    def writeExcel(self, write_src:str, data:list):
        curr_src = os.path.abspath("")  #获取当前脚本运行的绝对路径
        saveExcel = curr_src + "\\" + write_src
        sheet:worksheet = None
        excel:openpyxl.workbook = None
        try:
            #检测要写入的表格是否存在
            fs = open(write_src, "r")
            fs.close()
            self.__excel_op = openpyxl.load_workbook(write_src)  # 打开excel
            excel = self.__excel_op
            sheet = self.sheet
        except FileNotFoundError:
            print("不存在此表格文件，进行新建操作！")
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
            sheet = outws
            excel = outwb

        raw = 0
        sheet = excel.active  #激活追加写入
        for value in data:
            if value != None:
                raw += 1
                sheet.delete_rows(raw, 1)  #删除当前行
                sheet.append(value)
        excel.save(saveExcel)
        excel.close()

# table = excelProcess()
# table.readExcel(r"./test.xlsx", "Sheet1")
# line = table.getTableRowValue(3)
# colum = table.getTableColumValue(3)
# print(colum)